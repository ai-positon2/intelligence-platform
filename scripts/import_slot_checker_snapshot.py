"""Turn the Slot Checker .xlsx export into data/slot_checker_snapshot.json.

The Slot Checker agent writes its results to a Google Sheet that the platform's
service account cannot read: it is an office-internal sheet and Position2's
Workspace sharing policy blocks it (the same wall documented for the Job Change
Alert roster sheet, see app.py's JOB_CHANGE_TRACKED_SNAPSHOT_PATH). So the
dashboard reads a committed snapshot, and this script is how that snapshot is
regenerated from a fresh export. Run it, commit the JSON, redeploy.

    python3 scripts/import_slot_checker_snapshot.py ~/Downloads/"Slot Checker.xlsx"

Only two tabs are read, deliberately -- the workbook carries six near-duplicate
"Available Slots*" tabs and this is the one that is current:

  "All LPs"                -- the location registry, one row per practice
  "Available Slots Final"  -- one row per (practice, service), with one column
                              per date holding that day's open-slot count

Two things about the source that the parsing has to absorb rather than assume
away, both verified against the 2026-08-21 export:

  * The date columns are not a fixed window. They are whatever forward-looking
    range the agent happened to scrape, so they are read from the header row.
  * Neither the Location column nor the URL is a complete identity. The URL only
    carries location_title/state for the 46 practices that book through
    gentledental.com; the other 36 book through a Jarvis iframe keyed by an
    opaque location_id. The Location column is empty for 25 rows. Only
    "Location Name" (NNN-Brand-City-ST) is populated for all 82, so office
    number, brand and state come from there, and the display name prefers the
    Location column because it is the more specific of the two (it distinguishes
    Worcester-shrewsbury-st from Worcester-at-the-trolley-yard, which
    "Location Name" flattens to two rows both reading Worcester).
"""
from __future__ import annotations

import datetime
import json
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
OUT_PATH = REPO / "data" / "slot_checker_snapshot.json"

LP_TAB = "All LPs"
SLOT_TAB = "Available Slots Final"

# NNN-Brand-City-ST, where -ST is genuinely optional: 079-GentleDental-Dover has
# no state suffix, and its state (NH) is only recoverable from its URL.
_LOCATION_NAME = re.compile(r"^\s*(\d+)\s*-\s*([A-Za-z0-9]+)\s*-\s*(.+?)(?:-([A-Z]{2}))?\s*$")

# Words that stay lowercase when a hyphenated slug is turned into a display name.
_MINOR = {"at", "the", "of", "on", "and"}


def _clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    return str(v).strip()


def _iso_date(v) -> str:
    """A header cell as YYYY-MM-DD, or '' if it is not a date at all."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date().isoformat() if isinstance(v, datetime.datetime) else v.isoformat()
    s = _clean(v)
    m = re.match(r"^(\d{4}-\d{2}-\d{2})", s)
    if m:
        return m.group(1)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def _iso_stamp(v) -> str:
    if isinstance(v, datetime.datetime):
        return v.isoformat()
    s = _clean(v)
    try:
        return datetime.datetime.fromisoformat(s).isoformat()
    except ValueError:
        return s


def _titleise(raw: str) -> str:
    """'worcester-at-the-trolley-yard' -> 'Worcester at the Trolley Yard'."""
    parts = [p for p in re.split(r"[-_\s]+", raw.strip()) if p]
    out = []
    for i, p in enumerate(parts):
        low = p.lower()
        if i and low in _MINOR:
            out.append(low)
        elif p.isupper() and len(p) <= 3:
            out.append(p)                     # keep MA / NH / CT / RADA-ish tokens
        else:
            out.append(p[:1].upper() + p[1:])
        # 'Dover,NH' -> the comma-state form the Location column sometimes uses
    return " ".join(out)


def _split_camel(raw: str) -> str:
    """'WorcesterTrolley' -> 'Worcester Trolley'; leaves 'RADA' alone."""
    if raw.isupper():
        return raw
    return re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", raw)


def _url_bits(url: str) -> tuple[str, str, str]:
    """(system, key, state) from a booking URL.

    gentledental.com carries location_title + state; the Jarvis iframe carries an
    opaque location_id and no state at all.
    """
    u = url or ""
    m = re.search(r"location_title=([^&]+)", u)
    if m:
        st = re.search(r"[?&]state=([A-Za-z]{2})", u)
        return "gentledental", m.group(1).lower(), (st.group(1).upper() if st else "")
    m = re.search(r"location_id=(\d+)", u)
    if m:
        return "jarvis", m.group(1), ""
    return "other", u, ""


def parse_workbook(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    for tab in (LP_TAB, SLOT_TAB):
        if tab not in wb.sheetnames:
            raise SystemExit(f"{path.name} has no {tab!r} tab (found: {wb.sheetnames})")

    return build_snapshot(
        [list(r) for r in wb[LP_TAB].iter_rows(values_only=True)],
        [list(r) for r in wb[SLOT_TAB].iter_rows(values_only=True)],
        source=path.name,
    )


def build_snapshot(lp_rows: list, slot_rows: list, source: str = "") -> dict:
    """The whole parse, as a pure function over two sheets-shaped row lists.

    Taking rows rather than a file is what lets the tests exercise the awkward
    real-world shapes (missing state, empty Location, duplicate city names, a
    practice with no slot rows at all) without building an .xlsx each time, and
    is also the seam a live Sheets read would plug into unchanged.
    """
    dates, by_url = _parse_slots(slot_rows)
    locations = _parse_locations(lp_rows, dates, by_url)
    return {
        "generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "source": {"file": source, "tabs": [LP_TAB, SLOT_TAB]},
        "dates": dates,
        "locations": locations,
    }


def _parse_slots(rows: list) -> tuple[list, dict]:
    """-> (date list, {url: {service: [observation, ...]}}), oldest first.

    A single export holds MORE THAN ONE RUN. In the 2026-08-21 export, 111 of the
    226 (practice, service) pairs appear 2-4 times, once per run on 08-11, 08-12
    and 08-13, each with different counts because real availability moved between
    runs. So a row is an observation, not a fact, and summing the sheet naively
    triple-counts the practices that happen to have been re-scraped. Observations
    are kept per (practice, service) and the newest one is what "current" means.
    """
    if not rows:
        return [], {}
    header = rows[0]
    # Columns 0-2 are Url / Service Name / Execution Time; everything after is a
    # date, and trailing blank columns are openpyxl padding, not real columns.
    dates, date_cols = [], []
    for i, cell in enumerate(header[3:], start=3):
        iso = _iso_date(cell)
        if iso:
            dates.append(iso)
            date_cols.append(i)

    by_url: dict = {}
    for row in rows[1:]:
        url = _clean(row[0] if len(row) > 0 else "")
        service = _clean(row[1] if len(row) > 1 else "")
        if not url or not service:
            continue
        counts = []
        for i in date_cols:
            v = row[i] if len(row) > i else None
            try:
                counts.append(max(0, int(float(v))) if v not in (None, "") else 0)
            except (TypeError, ValueError):
                counts.append(0)
        obs = {"at": _iso_stamp(row[2] if len(row) > 2 else ""), "counts": counts}
        by_url.setdefault(url, {}).setdefault(service, []).append(obs)

    for services in by_url.values():
        for obs_list in services.values():
            obs_list.sort(key=lambda o: o["at"])
    return dates, by_url


def _parse_locations(rows: list, dates: list, by_url: dict) -> list:
    out = []
    for row in rows[1:] if rows else []:
        account = _clean(row[0] if len(row) > 0 else "")
        location = _clean(row[1] if len(row) > 1 else "")
        url = _clean(row[2] if len(row) > 2 else "")
        booking = _clean(row[3] if len(row) > 3 else "")
        loc_name = _clean(row[4] if len(row) > 4 else "")
        if not url and not loc_name:
            continue

        office = brand = city = state = ""
        m = _LOCATION_NAME.match(loc_name)
        if m:
            office, brand, city, state = m.group(1), m.group(2), m.group(3), (m.group(4) or "")

        system, key, url_state = _url_bits(url)
        # The URL is the only place Dover's state survives.
        state = state or url_state

        # Prefer the Location column: it is the more specific of the two, which
        # is what keeps the two Worcester practices distinguishable.
        if location:
            display = _titleise(location.replace(",", " "))
        elif city:
            display = _titleise(_split_camel(city))
        else:
            display = account or key

        services = [
            {"name": name, "observations": obs}
            for name, obs in sorted(by_url.get(url, {}).items())
        ]
        stamps = [o["at"] for s in services for o in s["observations"] if o["at"]]
        out.append({
            "office": office,
            "brand": _split_camel(brand) if brand else "",
            "account": account,
            "name": display,
            "city": _titleise(_split_camel(city)) if city else display,
            "state": state,
            "url": url,
            "system": system,
            "key": key,
            "booking": booking,
            "checked_at": max(stamps) if stamps else "",
            "first_checked_at": min(stamps) if stamps else "",
            "services": services,
        })
    out.sort(key=lambda r: (r["state"], r["name"]))
    return out


def main(argv: list) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 2
    src = Path(argv[1]).expanduser()
    if not src.exists():
        print(f"no such file: {src}")
        return 1
    snap = parse_workbook(src)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(snap, separators=(",", ":"), sort_keys=True))
    n_svc = sum(len(l["services"]) for l in snap["locations"])
    n_obs = sum(len(sv["observations"]) for l in snap["locations"] for sv in l["services"])
    print(json.dumps({
        "written": str(OUT_PATH.relative_to(REPO)),
        "locations": len(snap["locations"]),
        "service_pairs": n_svc,
        "observations": n_obs,
        "dates": len(snap["dates"]),
        "window": [snap["dates"][0], snap["dates"][-1]] if snap["dates"] else [],
        "bytes": OUT_PATH.stat().st_size,
    }))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
