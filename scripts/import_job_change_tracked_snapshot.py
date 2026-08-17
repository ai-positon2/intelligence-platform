"""Import the tracked-contacts/tracked-companies roster from a manual .xlsx
export of the "Job_change_tracker-For SFO Companies" Google Sheet into
data/job_change_tracked_snapshot.json.

Why this exists: the platform's Google service account cannot read that
sheet directly -- a Position2 Workspace external-sharing policy blocks
sharing it with a non-allowlisted service account (confirmed 2026-08-17;
see the project_job_change_alert memory for the exact error). Until a
Workspace admin allowlists it (or the sheet is re-shared from a personal
account), _fetch_job_change_tracked_data() in app.py falls back to this
committed snapshot whenever the live Sheets read comes back empty.

Usage:
    python3 scripts/import_job_change_tracked_snapshot.py path/to/export.xlsx

Re-run this against a fresh export whenever the roster sheet changes and
commit the updated data/job_change_tracked_snapshot.json.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

CONTACTS_HEADER_ROW = 5   # matches the live sheet's "Contact List (Being Monitored)" tab
COMPANIES_HEADER_ROW = 1  # matches the live sheet's "Tracked Companies" tab


def _to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _header_map(row) -> dict:
    return {str(h).strip().lower(): i for i, h in enumerate(row) if h and str(h).strip()}


def _col(row, idx, default: str = "") -> str:
    if idx is None or idx >= len(row) or row[idx] is None:
        return default
    return _to_str(row[idx])


def import_snapshot(xlsx_path: str) -> dict:
    import openpyxl  # local import: only needed for this one-off script

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    ws = wb["Contact List (Being Monitored)"]
    rows = [r for r in ws.iter_rows(min_row=CONTACTS_HEADER_ROW, values_only=True)]
    hdr = _header_map(rows[0])
    contacts = []
    for r in rows[1:]:
        name = (_col(r, hdr.get("first name")) + " " + _col(r, hdr.get("last name"))).strip()
        if not name:
            continue
        contacts.append({
            "name": name,
            "title": _col(r, hdr.get("title")),
            "company": _col(r, hdr.get("company name")),
            "seniority": _col(r, hdr.get("seniority")),
            "department": _col(r, hdr.get("departments")),
            "industry": _col(r, hdr.get("industry")),
            "employees": _col(r, hdr.get("# employees")),
            "city": _col(r, hdr.get("city")),
            "state": _col(r, hdr.get("state")),
            "linkedin_url": _col(r, hdr.get("person linkedin url")),
            "company_linkedin_url": _col(r, hdr.get("company linkedin url")),
            "website": _col(r, hdr.get("website")),
        })

    ws2 = wb["Tracked Companies"]
    rows2 = [r for r in ws2.iter_rows(min_row=COMPANIES_HEADER_ROW, values_only=True)]
    hdr2 = _header_map(rows2[0])
    companies = []
    for r in rows2[1:]:
        name = _col(r, hdr2.get("company name"))
        if not name:
            continue
        companies.append({
            "name": name,
            "industry": _col(r, hdr2.get("industry")),
            "employees": _col(r, hdr2.get("# employees")),
            "website": _col(r, hdr2.get("website")),
            "linkedin_url": _col(r, hdr2.get("company linkedin url")),
            "city": _col(r, hdr2.get("city")),
            "state": _col(r, hdr2.get("state")),
            "country": _col(r, hdr2.get("country")),
            "annual_revenue": _col(r, hdr2.get("annual revenue")),
            "total_funding": _col(r, hdr2.get("total funding")),
            "latest_funding": _col(r, hdr2.get("latest funding")),
        })

    contacts.sort(key=lambda c: c["name"])
    companies.sort(key=lambda c: c["name"])
    return {"contacts": contacts, "companies": companies}


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: python3 scripts/import_job_change_tracked_snapshot.py path/to/export.xlsx")
        sys.exit(1)

    result = import_snapshot(sys.argv[1])
    out_path = Path(__file__).parent.parent / "data" / "job_change_tracked_snapshot.json"
    payload = {
        "_readme": (
            "Snapshot of the tracked-contacts/tracked-companies roster from the "
            "'Job_change_tracker-For SFO Companies' Google Sheet, imported from a manual "
            ".xlsx export because the sheet cannot be shared with the platform's Google "
            "service account (blocked by a Position2 Workspace external-sharing policy -- "
            "see the project_job_change_alert memory). Used as a fallback by "
            "_fetch_job_change_tracked_data() in app.py whenever the live Sheets read "
            "returns empty. Re-run this script against a fresh export to refresh."
        ),
        **result,
    }
    out_path.write_text(json.dumps(payload, indent=2))
    print(f"wrote {out_path}: {len(result['contacts'])} contacts, {len(result['companies'])} companies")
