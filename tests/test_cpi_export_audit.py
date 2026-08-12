"""The export flow, audited the way the filters, the chat and enrich were.

Export is the only surface whose output outlives the session that made it. A
spreadsheet gets filtered, sorted, mail-merged and pasted into a CRM months later
by someone who never saw the screen it came from, so anything the screen explains
with a badge, a tooltip or a prompt has to be a column instead. That is the whole
finding here, in four places.

Checked and already correct, so only pinned:

  Formula injection is defused on both the csv and the xlsx path, and phone
  numbers and negative figures are deliberately left alone rather than picking up
  a stray apostrophe.

  Export costs no Apollo credits. Rows come from the client, so nothing is
  re-queried.

  Only verified rows are exported: the client holds what survived the checks.

  The reported total is not stale. STATE.total resets to null per search, and the
  server sends null whenever verification rejected anything, so the sheet omits
  the line rather than printing a figure from an earlier query.

Four mismatches found and fixed:

  1. A masked surname exported as the name. Apollo withholds surnames as
     asterisks; on screen that comes with a "masked" badge and a tooltip saying
     enrichment reveals it, and in the file "Vivek Sh***a" sat under a header
     called Name with nothing to say it was incomplete.
  2. An empty Email cell meant either "nobody has spent a credit on this person"
     or "a credit was spent and Apollo holds no address". The card distinguishes
     them with an Enrich prompt; the file could not.
  3. The "Search details" sheet printed Apollo organization ids as 24-character
     hex. A company-scoped search is the commonest scoped search there is, and the
     sheet answered "which company" with an opaque id.
  4. Auto-generated labels mangled the ones that are not plain words: "Naics
     codes", "Sic codes", and "Employee min" for a lower bound.

And one thing the sheet did not say at all: how many rows Apollo offered that our
own checks then removed. Without it the file reads as everything Apollo returned,
when it is deliberately less, and that difference is the point of the whole
verification pass.
"""

import os
import sys

import pytest

os.environ.setdefault("GOOGLE_CLIENT_ID", "test")
os.environ.setdefault("GOOGLE_CLIENT_SECRET", "test")
os.environ.setdefault("FLASK_SECRET_KEY", "test")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import app as appmod  # noqa: E402

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_JS = os.path.join(_ROOT, "static", "js", "company_people_intelligence.js")

_EXPORT = "/p2/b2b-agents/company-people-intelligence/export"


@pytest.fixture
def client():
    c = appmod.app.test_client()
    with c.session_transaction() as sess:
        sess["google_user"] = {"email": "reporting@position2.com",
                               "name": "Test User", "given_name": "Test"}
    return c


_MASKED = {"id": "m1", "full_name": "Vivek Sh***a", "name_masked": True,
           "title": "VP Engineering", "organization_name": "Acme"}
_ENRICHED = {"id": "e1", "full_name": "Jane Doe", "name_masked": False,
             "title": "CMO", "email": "jane@acme.com", "email_status": "verified",
             "enriched": True, "organization_name": "Acme"}
_PLAIN = {"id": "p1", "full_name": "Sam Roe", "name_masked": False,
          "title": "CFO", "organization_name": "Acme"}


def _csv(client, rows, **body):
    payload = {"entity": "people", "format": "csv", "rows": rows}
    payload.update(body)
    r = client.post(_EXPORT, json=payload)
    assert r.status_code == 200
    lines = r.data.decode("utf-8-sig").splitlines()
    import csv as _c
    return list(_c.reader(lines))


def _col(table, label):
    """The values under one header, by label, so a test never depends on order."""
    idx = table[0].index(label)
    return [row[idx] for row in table[1:]]


def _js():
    return open(_JS, encoding="utf-8").read()


# ── The masked surname ────────────────────────────────────────────────────────

def test_a_withheld_surname_is_flagged_in_its_own_column(client):
    """A spreadsheet has no tooltips. The badge the screen shows has to be a cell."""
    table = _csv(client, [_MASKED, _PLAIN])
    flags = _col(table, "Surname withheld by Apollo")
    assert flags[0].startswith("Yes")
    assert flags[1] == ""


def test_the_masked_name_itself_is_still_exported_as_apollo_returned_it(client):
    """Not replaced with a prettified "Vivek Sh.": the asterisked form still tells
    two people with the same first name apart, and shortening it would lose that
    while staying just as incomplete."""
    table = _csv(client, [_MASKED])
    assert _col(table, "Name") == ["Vivek Sh***a"]


def test_a_row_saved_before_the_flag_existed_is_still_flagged(client):
    """History entries predate name_masked, so the asterisks are treated as proof
    on their own rather than trusting a field that may not be there."""
    table = _csv(client, [{"id": "x", "full_name": "Vivek Sh***a", "title": "VP"}])
    assert _col(table, "Surname withheld by Apollo")[0].startswith("Yes")


def test_a_name_with_no_asterisks_and_no_flag_is_left_alone(client):
    table = _csv(client, [{"id": "x", "full_name": "Ada Lovelace"}])
    assert _col(table, "Surname withheld by Apollo") == [""]


# ── An empty contact cell ─────────────────────────────────────────────────────

def test_an_empty_email_says_which_kind_of_empty_it_is(client):
    """Never enriched, versus enriched and Apollo has nothing. The two look
    identical in a spreadsheet and mean opposite things about whether spending a
    credit would help."""
    table = _csv(client, [_ENRICHED, _PLAIN])
    revealed = _col(table, "Contact details revealed")
    assert revealed[0] == "Yes"
    assert revealed[1].startswith("No")


def test_a_person_enriched_to_a_blank_email_still_counts_as_revealed(client):
    """The credit was spent. Saying "not enriched" would invite spending it twice."""
    table = _csv(client, [{"id": "z", "full_name": "No Email Person",
                           "enriched": True}])
    assert _col(table, "Contact details revealed") == ["Yes"]


def test_a_row_with_an_email_but_no_flag_counts_as_revealed(client):
    """Older history rows carry the email without the flag."""
    table = _csv(client, [{"id": "z", "full_name": "A B", "email": "a@b.com"}])
    assert _col(table, "Contact details revealed") == ["Yes"]


# ── The search details sheet ───────────────────────────────────────────────────

def test_an_organization_id_is_never_printed_as_hex():
    """The sheet exists to say which search produced these rows. Answering "which
    company" with 5e66b6381e05b4008c8331b8 does not."""
    out = dict(appmod._cpi_filters_readable(
        {"organization_ids": ["5e66b6381e05b4008c8331b8"]}))
    value = out["Scoped to specific companies"]
    assert "5e66b6381e05b4008c8331b8" not in value
    assert "1 company" in value


def test_several_scoped_companies_are_counted_in_the_plural():
    out = dict(appmod._cpi_filters_readable(
        {"organization_ids": ["a" * 24, "b" * 24, "c" * 24]}))
    assert "3 companies" in out["Scoped to specific companies"]


@pytest.mark.parametrize("key,label", [
    ("naics_codes", "NAICS codes"),
    ("sic_codes", "SIC codes"),
    ("exclude_naics_codes", "NAICS codes excluded"),
    ("employee_min", "Employees from"),
    ("employee_max", "Employees up to"),
    ("person_locations", "Person location"),
    ("company_locations", "Employer HQ location"),
    ("exclude_technologies", "Does not use these technologies"),
])
def test_the_labels_read_as_english_not_as_field_names(key, label):
    out = dict(appmod._cpi_filters_readable({key: ["5415"] if "codes" in key else 5}))
    assert label in out


def test_a_key_with_no_special_label_still_gets_a_readable_one():
    """The map only covers the ones capitalize() gets wrong; everything else must
    keep working without an entry."""
    out = dict(appmod._cpi_filters_readable({"titles": ["CMO"]}))
    assert out["Titles"] == "CMO"


def test_the_fetch_toggle_is_labelled_as_a_fetch_not_a_filter():
    """It is not a constraint on which rows came back. It IS the reason employer
    columns are blank in one file and full in another, which is a question a reader
    of an old spreadsheet genuinely asks, so it stays under an honest label rather
    than being dropped."""
    out = dict(appmod._cpi_filters_readable({"company_detail": False}))
    assert out["Employer details fetched"] == "No"
    assert "Company detail" not in out


def test_internal_page_caps_are_not_listed_as_search_constraints():
    out = dict(appmod._cpi_filters_readable(
        {"titles": ["CMO"], "max_people": 25, "max_companies": 10}))
    assert list(out) == ["Titles"]


def test_rows_removed_on_checking_are_recorded_in_the_file(client):
    """Otherwise the file reads as everything Apollo offered, when it is
    deliberately less, and that gap is the point of the verification pass."""
    r = client.post(_EXPORT, json={
        "entity": "people", "format": "xlsx", "rows": [_PLAIN],
        "filters": {"titles": ["CMO"]},
        "meta": {"total": 40, "rejected": {"industry": 6, "employees": 2}}})
    assert r.status_code == 200
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(r.data))
    sheet = wb["Search details"]
    pairs = {row[0]: row[1] for row in sheet.iter_rows(min_row=2, values_only=True)}
    labels = " ".join(str(k) for k in pairs)
    assert "Removed on checking" in labels
    # Named by reason, in the same words the results header uses, and biggest first.
    assert pairs["Removed on checking: outside the industry"] == "6"
    assert pairs["Removed on checking: outside the size range"] == "2"
    order = [k for k in pairs if str(k).startswith("Removed on checking")]
    assert order[0].endswith("outside the industry")


def test_a_clean_page_says_nothing_about_removals(client):
    """A note that appears on every export stops carrying information."""
    r = client.post(_EXPORT, json={
        "entity": "people", "format": "xlsx", "rows": [_PLAIN],
        "filters": {"titles": ["CMO"]}, "meta": {"total": 1, "rejected": {}}})
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(r.data))
    labels = " ".join(str(row[0]) for row in
                      wb["Search details"].iter_rows(min_row=2, values_only=True))
    assert "Removed on checking" not in labels


def test_a_zero_count_is_not_reported_as_a_removal(client):
    r = client.post(_EXPORT, json={
        "entity": "people", "format": "xlsx", "rows": [_PLAIN],
        "filters": {"titles": ["CMO"]}, "meta": {"rejected": {"industry": 0}}})
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(r.data))
    labels = " ".join(str(row[0]) for row in
                      wb["Search details"].iter_rows(min_row=2, values_only=True))
    assert "Removed on checking" not in labels


def test_the_client_sends_the_removal_counts_it_already_holds():
    js = _js()
    block = js[js.index("window.cpiExport"):]
    block = block[:block.index("\n};")]
    assert "rejected: STATE.rejected" in block
    # A hand-picked selection is not the results of the search any more, so it may
    # not claim the search's bookkeeping either.
    assert "onlySelected ? {} :" in block


# ── Things that were already right, pinned so they stay right ──────────────────

@pytest.mark.parametrize("value,expected", [
    ("=cmd|'/c calc'!A1", "'=cmd|'/c calc'!A1"),
    ("@SUM(A1:A9)", "'@SUM(A1:A9)"),
    ("=HYPERLINK(\"http://x\")", "'=HYPERLINK(\"http://x\")"),
    # Left alone: a quoted phone number shows a stray apostrophe in every row.
    ("+1 (555) 123-4567", "+1 (555) 123-4567"),
    ("-42", "-42"),
    ("Acme Inc", "Acme Inc"),
])
def test_formula_injection_is_defused_without_mangling_phone_numbers(value, expected):
    assert appmod._csv_safe(value) == expected


def test_the_xlsx_path_defuses_formulas_too(client):
    """A dangerous cell is dangerous in both formats, and the xlsx path builds its
    cells separately."""
    r = client.post(_EXPORT, json={
        "entity": "companies", "format": "xlsx",
        "rows": [{"name": "=cmd|'/c calc'!A1", "primary_domain": "x.com"}]})
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(r.data))
    assert wb.active.cell(row=2, column=1).value.startswith("'=")


def test_the_xlsx_carries_the_derived_columns_too(client):
    """The two formats build their cells through separate loops, and xlsx is the
    one people actually keep. A fix applied to only one of them leaves the
    long-lived format as the dishonest one."""
    r = client.post(_EXPORT, json={"entity": "people", "format": "xlsx",
                                   "rows": [_MASKED, _ENRICHED]})
    assert r.status_code == 200
    from openpyxl import load_workbook
    import io
    ws = load_workbook(io.BytesIO(r.data))["People"]
    header = [c.value for c in ws[1]]
    masked_col = header.index("Surname withheld by Apollo") + 1
    contact_col = header.index("Contact details revealed") + 1
    assert str(ws.cell(row=2, column=masked_col).value or "").startswith("Yes")
    assert str(ws.cell(row=2, column=contact_col).value or "").startswith("No")
    assert ws.cell(row=3, column=masked_col).value in (None, "")
    assert ws.cell(row=3, column=contact_col).value == "Yes"


def test_exporting_costs_no_apollo_credits(client, monkeypatch):
    """Rows come from the client precisely so that downloading what you already
    have is free."""
    import tracker.apollo_client as ac
    for name in ("_post", "search_people", "search_companies", "bulk_match_people"):
        monkeypatch.setattr(ac, name,
                            lambda *a, **k: pytest.fail("export must not call Apollo"))
    r = client.post(_EXPORT, json={"entity": "people", "format": "csv",
                                   "rows": [_PLAIN]})
    assert r.status_code == 200


def test_an_empty_export_is_refused_rather_than_returning_an_empty_file(client):
    r = client.post(_EXPORT, json={"entity": "people", "format": "csv", "rows": []})
    assert r.status_code == 400


def test_the_csv_stays_a_flat_table(client):
    """No second sheet to hold context, so nothing about the search may be appended
    as extra lines that break a re-import."""
    table = _csv(client, [_PLAIN], filters={"titles": ["CMO"]},
                 meta={"total": 42, "rejected": {"industry": 3}})
    assert len(table) == 2
    assert "42" not in ",".join(table[0])


def test_every_person_column_has_a_value_the_exporter_can_produce(client):
    """A column whose key nothing produces is a permanently empty column, which
    reads as "Apollo has no data" rather than "this was never wired up"."""
    table = _csv(client, [_ENRICHED])
    assert len(table[0]) == len(appmod._CPI_PERSON_COLS)
    assert len(table[1]) == len(appmod._CPI_PERSON_COLS)


def test_the_derived_columns_are_labelled_as_derived():
    """Established discipline on this page: a column called Seniority holding a
    value Apollo never asserted is a quiet fiction a spreadsheet carries forever."""
    labels = [lbl for _k, lbl in appmod._CPI_PERSON_COLS]
    assert "Seniority (from title)" in labels
    assert "Function (from title)" in labels
    # And Apollo's own value keeps the plain name, so the two are distinguishable.
    assert "Seniority" in labels
