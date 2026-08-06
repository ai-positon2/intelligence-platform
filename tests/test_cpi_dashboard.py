"""Tests for the Contact Finder dashboard: bulk enrich, history,
export, and the OpenAI model/reasoning ladder.

The dashboard's job is to show as much as Apollo gives away for free and to make
every credit-spending step explicit, so these lean on two themes:
  1. Nothing silently spends Apollo credits (bulk enrich is capped, cached ids
     are reported separately, export re-uses rows the client already has).
  2. Nothing silently degrades (a bad model id or an unsupported reasoning
     parameter must fall through, not surface as a failed chat).
"""

import os
import sys
import types

import pytest

os.environ.setdefault("GOOGLE_CLIENT_ID", "test")
os.environ.setdefault("GOOGLE_CLIENT_SECRET", "test")
os.environ.setdefault("FLASK_SECRET_KEY", "test")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import app as appmod  # noqa: E402


@pytest.fixture
def client():
    c = appmod.app.test_client()
    with c.session_transaction() as sess:
        sess["google_user"] = {"email": "reporting@position2.com",
                               "name": "Test User", "given_name": "Test"}
    return c


@pytest.fixture
def no_postgres(monkeypatch):
    """History and the id cache must degrade gracefully with no database."""
    monkeypatch.setattr(appmod, "_pg_conn", lambda: None)


# ── Spreadsheet-formula injection ────────────────────────────────────────────

@pytest.mark.parametrize("raw", [
    "=cmd|' /c calc'!A1",
    '+HYPERLINK("http://evil","click")',
    "@SUM(A1:A9)",
    "-2+3+cmd|' /c calc'!A0",
])
def test_csv_safe_defuses_formulas(raw):
    """Apollo text lands in files people open in Excel, so a leading formula
    character has to be made inert -- with the text still readable."""
    out = appmod._csv_safe(raw)
    assert out.startswith("'")
    assert out[1:] == raw


@pytest.mark.parametrize("raw", ["+1 555 0100", "+91 (80) 4718-1000", "-4200000", "Acme Corp"])
def test_csv_safe_leaves_real_values_alone(raw):
    """Phone numbers start with "+" and revenue can be negative; quoting those
    would put a stray apostrophe in every phone and currency column."""
    assert appmod._csv_safe(raw) == raw


def test_csv_safe_flattens_lists_and_blanks():
    assert appmod._csv_safe(["a", "", None, "b"]) == "a, b"
    assert appmod._csv_safe(None) == ""
    assert appmod._csv_safe(False) == ""


# ── Export ───────────────────────────────────────────────────────────────────

_ROW = {"full_name": "Ada Lovelace", "title": "CMO", "email": "ada@acme.com",
        "organization_name": "Acme", "departments": ["marketing"],
        "organization_employees": 240, "id": "p1"}


def test_export_csv_has_headers_and_bom(client):
    r = client.post("/p2/gtm/company-people-intelligence/export",
                    json={"entity": "people", "format": "csv", "rows": [_ROW]})
    assert r.status_code == 200
    assert r.headers["Content-Disposition"].startswith('attachment; filename="apollo-people-')
    # utf-8-sig BOM, so Excel reads non-ASCII company names correctly.
    assert r.data.startswith(b"\xef\xbb\xbf")
    body = r.data.decode("utf-8-sig")
    assert body.splitlines()[0].startswith("Name,Title,Seniority,Email")
    assert "Ada Lovelace" in body and "marketing" in body


def test_export_xlsx_is_a_real_workbook(client):
    import io
    import openpyxl
    r = client.post("/p2/gtm/company-people-intelligence/export",
                    json={"entity": "people", "format": "xlsx", "rows": [_ROW]})
    assert r.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(r.data)).active
    assert ws.title == "People"
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=2, column=1).value == "Ada Lovelace"
    assert ws.freeze_panes == "A2"


def test_export_companies_uses_company_columns(client):
    r = client.post("/p2/gtm/company-people-intelligence/export",
                    json={"entity": "companies", "format": "csv",
                          "rows": [{"name": "Acme", "primary_domain": "acme.com",
                                    "technologies": ["salesforce", "hubspot"]}]})
    body = r.data.decode("utf-8-sig")
    assert body.splitlines()[0].startswith("Company,Domain,Industry")
    assert "salesforce, hubspot" in body


def test_export_rejects_empty_selection(client):
    r = client.post("/p2/gtm/company-people-intelligence/export",
                    json={"entity": "people", "rows": []})
    assert r.status_code == 400


def test_export_defaults_to_xlsx_for_unknown_format(client):
    r = client.post("/p2/gtm/company-people-intelligence/export",
                    json={"entity": "people", "format": "exe", "rows": [_ROW]})
    assert r.status_code == 200
    assert r.headers["Content-Disposition"].endswith('.xlsx"')


# ── Bulk enrich ──────────────────────────────────────────────────────────────

def test_bulk_enrich_is_capped_and_deduped(client, monkeypatch, no_postgres):
    """An uncapped 'enrich all' over several pages could drain a shared pool, so
    the route must never ask Apollo for more than the cap, even if asked to."""
    seen = {}

    def _fake_bulk(ids, api_key):
        seen["ids"] = ids
        return {i: {"id": i, "first_name": "X", "last_name": "Y"} for i in ids}

    monkeypatch.setenv("APOLLO_API_KEY", "k")
    import tracker.apollo_client as ac
    monkeypatch.setattr(ac, "bulk_match_people", _fake_bulk)

    # 120 ids, each duplicated once.
    ids = [f"p{i}" for i in range(120)] * 2
    r = client.post("/p2/gtm/company-people-intelligence/enrich-bulk", json={"ids": ids})
    assert r.status_code == 200
    body = r.get_json()
    assert len(seen["ids"]) == appmod._CPI_BULK_ENRICH_CAP
    assert len(set(seen["ids"])) == len(seen["ids"]), "duplicate ids would double-charge"
    assert body["capped"] is True
    assert body["fetched"] == appmod._CPI_BULK_ENRICH_CAP


def test_bulk_enrich_reports_cache_hits_separately(client, monkeypatch):
    """Cached ids cost nothing, so the UI has to be able to say what a click
    actually spent rather than implying every row was billed."""
    monkeypatch.setenv("APOLLO_API_KEY", "k")
    monkeypatch.setattr(appmod, "_cpi_id_cache_read",
                        lambda ids: {"p1": {"id": "p1", "first_name": "Cached"}})
    monkeypatch.setattr(appmod, "_cpi_id_cache_write", lambda profiles: None)
    import tracker.apollo_client as ac
    monkeypatch.setattr(ac, "bulk_match_people",
                        lambda ids, key: {"p2": {"id": "p2", "first_name": "Fresh"}})

    body = client.post("/p2/gtm/company-people-intelligence/enrich-bulk",
                       json={"ids": ["p1", "p2"]}).get_json()
    assert body["cached"] == 1
    assert body["fetched"] == 1
    assert set(body["profiles"]) == {"p1", "p2"}


def test_bulk_enrich_empty_input_never_calls_apollo(client, monkeypatch):
    called = []
    import tracker.apollo_client as ac
    monkeypatch.setattr(ac, "bulk_match_people",
                        lambda ids, key: called.append(ids) or {})
    body = client.post("/p2/gtm/company-people-intelligence/enrich-bulk",
                       json={"ids": ["", "  ", None]}).get_json()
    assert body == {"profiles": {}, "fetched": 0, "cached": 0}
    assert not called


def test_bulk_enrich_survives_apollo_failure_with_cached_rows(client, monkeypatch):
    """A transport failure must still return whatever was already cached."""
    monkeypatch.setenv("APOLLO_API_KEY", "k")
    monkeypatch.setattr(appmod, "_cpi_id_cache_read",
                        lambda ids: {"p1": {"id": "p1", "first_name": "Cached"}})
    import tracker.apollo_client as ac

    def _boom(ids, key):
        raise RuntimeError("apollo down")
    monkeypatch.setattr(ac, "bulk_match_people", _boom)

    body = client.post("/p2/gtm/company-people-intelligence/enrich-bulk",
                       json={"ids": ["p1", "p2"]}).get_json()
    assert "error" not in body
    assert set(body["profiles"]) == {"p1"}


def test_bulk_enrich_without_apollo_key_is_explicit(client, monkeypatch):
    monkeypatch.delenv("APOLLO_API_KEY", raising=False)
    body = client.post("/p2/gtm/company-people-intelligence/enrich-bulk",
                       json={"ids": ["p1"]}).get_json()
    assert "not configured" in body["error"]


def test_person_row_keeps_contact_fields_and_flags_enriched():
    row = appmod._cpi_person_row({
        "id": "p1", "first_name": "Ada", "last_name": "Lovelace", "title": "CMO",
        "email": "ada@acme.com", "email_status": "verified",
        "phone_numbers": [{"sanitized_number": "+15550100"}],
        "employment_history": [{"organization_name": "Globex", "current": False}],
        "organization": {"name": "Acme", "primary_domain": "acme.com"},
    })
    assert row["full_name"] == "Ada Lovelace"
    assert row["email"] == "ada@acme.com"
    assert row["phones"] == ["+15550100"]
    assert row["past_companies"] == ["Globex"]
    assert row["enriched"] is True
    assert row["name_masked"] is False


# ── History ──────────────────────────────────────────────────────────────────

def test_history_degrades_without_postgres(client, no_postgres):
    body = client.get("/p2/gtm/company-people-intelligence/history").get_json()
    assert body == {"entries": [], "available": False}


def test_history_entry_404s_without_postgres(client, no_postgres):
    assert client.get("/p2/gtm/company-people-intelligence/history/1").status_code == 404


@pytest.mark.parametrize("entity,filters,expected", [
    ("people", {"titles": ["CMO"], "person_locations": ["United States"]},
     "CMO · United States"),
    ("companies", {"name": "Acme"}, "Acme"),
    ("people", {"seniorities": ["c_suite"], "keywords": "fintech"},
     "c_suite · fintech"),
    ("people", {}, "All people"),
    ("companies", {}, "All companies"),
])
def test_history_label_summarises_the_search(entity, filters, expected):
    assert appmod._cpi_history_label(entity, filters) == expected


def test_history_label_is_length_bounded():
    label = appmod._cpi_history_label("people", {"titles": ["x" * 400]})
    assert len(label) <= 160


# ── Company row shaping ──────────────────────────────────────────────────────

def test_company_row_exposes_firmographics_and_bounds_description():
    row = appmod._cpi_company_row({
        "id": "o1", "name": "Acme", "primary_domain": "acme.com",
        "industry": "software", "estimated_num_employees": 240,
        "annual_revenue": 4_200_000, "total_funding": 9_000_000,
        "founded_year": 2015, "short_description": "d" * 900,
        "technology_names": ["a"] * 30, "keywords": ["k"] * 30,
    })
    assert row["industry"] == "software"
    assert row["annual_revenue"] == 4_200_000
    assert len(row["short_description"]) == 280
    assert len(row["technologies"]) == 12
    assert len(row["keywords"]) == 10


def test_company_row_survives_a_sparse_record():
    row = appmod._cpi_company_row({"id": "o2", "name": "Tiny"})
    assert row["name"] == "Tiny"
    assert row["short_description"] is None
    assert row["technologies"] == []


# ── OpenAI model + reasoning ladder ──────────────────────────────────────────

def test_model_chain_leads_with_gpt56_then_55(monkeypatch):
    monkeypatch.delenv("OPENAI_INSIGHTS_MODEL", raising=False)
    chain = appmod._vimi_model_chain()
    assert chain[0] == "gpt-5.6-sol"
    assert chain[1] == "gpt-5.5"
    # Older ids stay on as a safety net so chat never hard-fails.
    assert "gpt-4o-mini" in chain


def test_insights_model_env_jumps_the_queue(monkeypatch):
    monkeypatch.setenv("OPENAI_INSIGHTS_MODEL", "gpt-6-preview")
    assert appmod._vimi_model_chain()[0] == "gpt-6-preview"


def test_reasoning_attempts_try_max_effort_first_with_headroom(monkeypatch):
    """Reasoning tokens bill against max_completion_tokens, so a 500-token ask
    must be raised or the model can spend the whole budget thinking and return
    empty content -- which looks like a model failure."""
    monkeypatch.delenv("OPENAI_REASONING_EFFORT", raising=False)
    attempts = appmod._vimi_attempts(500)
    assert attempts[0][0]["reasoning_effort"] == "max"
    assert attempts[0][1] >= appmod._VIMI_REASONING_FLOOR
    assert attempts[1][0]["reasoning_effort"] == "high"
    # Final attempt drops reasoning entirely and keeps the caller's budget.
    assert attempts[-1][0] == {} and attempts[-1][1] == 500


def test_temperature_only_rides_non_reasoning_attempts(monkeypatch):
    monkeypatch.delenv("OPENAI_REASONING_EFFORT", raising=False)
    for kw, _budget in appmod._vimi_attempts(500, temperature=0.1):
        assert not ("reasoning_effort" in kw and "temperature" in kw)
    assert any("temperature" in kw for kw, _b in appmod._vimi_attempts(500, 0.1))


def test_reasoning_can_be_switched_off_to_legacy_behaviour(monkeypatch):
    monkeypatch.setenv("OPENAI_REASONING_EFFORT", "off")
    assert [kw for kw, _b in appmod._vimi_attempts(500)] == [{}]
    assert [kw for kw, _b in appmod._vimi_attempts(500, 0.1)] == [{"temperature": 0.1}, {}]


def test_pinned_effort_still_falls_through(monkeypatch):
    """An unsupported pinned value must not wedge every AI feature in the app."""
    monkeypatch.setenv("OPENAI_REASONING_EFFORT", "ludicrous")
    efforts = [kw.get("reasoning_effort") for kw, _b in appmod._vimi_attempts(500)]
    assert efforts == ["ludicrous", None]


class _FakeOAI:
    """Minimal OpenAI stand-in that accepts only certain reasoning_effort values."""

    def __init__(self, accepted):
        self.accepted = accepted
        self.tried = []
        outer = self

        def create(model, messages, max_completion_tokens, **kw):
            outer.tried.append(kw.get("reasoning_effort"))
            if kw.get("reasoning_effort") not in outer.accepted:
                raise RuntimeError("unsupported reasoning_effort")
            msg = types.SimpleNamespace(content="ok")
            return types.SimpleNamespace(choices=[types.SimpleNamespace(message=msg)])

        self.chat = types.SimpleNamespace(
            completions=types.SimpleNamespace(create=create))


def test_unsupported_reasoning_falls_back_then_is_remembered(monkeypatch):
    """Without memoisation, a model that rejects reasoning_effort would burn two
    dead round-trips on every AI call in the app."""
    monkeypatch.delenv("OPENAI_REASONING_EFFORT", raising=False)
    appmod._VIMI_EFFORT_OK.clear()
    oai = _FakeOAI(accepted={None})

    txt, err = appmod._vimi_create(oai, "m", [], 500)
    assert txt == "ok"
    assert oai.tried == ["max", "high", None]

    oai.tried = []
    txt, _ = appmod._vimi_create(oai, "m", [], 500)
    assert txt == "ok"
    assert oai.tried == [None], "should go straight to the learned value"
    appmod._VIMI_EFFORT_OK.clear()


def test_max_effort_model_is_used_on_the_first_try(monkeypatch):
    monkeypatch.delenv("OPENAI_REASONING_EFFORT", raising=False)
    appmod._VIMI_EFFORT_OK.clear()
    oai = _FakeOAI(accepted={"max"})
    txt, _ = appmod._vimi_create(oai, "m2", [], 500)
    assert txt == "ok" and oai.tried == ["max"]
    appmod._VIMI_EFFORT_OK.clear()


def test_empty_content_is_not_treated_as_success(monkeypatch):
    """A reasoning model that spends its whole budget thinking returns empty
    content with no exception; that must fall through, not surface as an answer."""
    monkeypatch.setenv("OPENAI_REASONING_EFFORT", "off")
    appmod._VIMI_EFFORT_OK.clear()

    def create(model, messages, max_completion_tokens, **kw):
        msg = types.SimpleNamespace(content="   ")
        return types.SimpleNamespace(choices=[types.SimpleNamespace(message=msg)])

    oai = types.SimpleNamespace(
        chat=types.SimpleNamespace(completions=types.SimpleNamespace(create=create)))
    txt, _err = appmod._vimi_create(oai, "m", [], 500)
    assert txt is None
    appmod._VIMI_EFFORT_OK.clear()


# ── Dead-model memoisation ───────────────────────────────────────────────────

class _DeadModelOAI:
    """OpenAI stand-in where some model ids do not exist on the account."""

    def __init__(self, dead_models, status=404):
        self.dead_models, self.status, self.tried = dead_models, status, []
        outer = self

        class _Err(Exception):
            status_code = status

        def create(model, messages, max_completion_tokens, **kw):
            outer.tried.append(model)
            if model in outer.dead_models:
                raise _Err("The model `%s` does not exist" % model)
            msg = types.SimpleNamespace(content="ok")
            return types.SimpleNamespace(choices=[types.SimpleNamespace(message=msg)])

        self.chat = types.SimpleNamespace(
            completions=types.SimpleNamespace(create=create))


@pytest.fixture
def clean_model_memo(monkeypatch):
    monkeypatch.delenv("OPENAI_REASONING_EFFORT", raising=False)
    monkeypatch.delenv("OPENAI_INSIGHTS_MODEL", raising=False)
    appmod._VIMI_DEAD.clear()
    appmod._VIMI_EFFORT_OK.clear()
    yield
    appmod._VIMI_DEAD.clear()
    appmod._VIMI_EFFORT_OK.clear()


def test_a_nonexistent_model_is_retired_after_one_call(clean_model_memo):
    """A wrong id in _VIMI_MODELS otherwise costs a full failed effort-ladder walk
    on every AI call in the app, for the life of the process."""
    chain = appmod._vimi_model_chain()
    oai = _DeadModelOAI({chain[0]})

    appmod._vimi_completion(oai, [], 500)
    assert chain[0] in oai.tried, "first call should still try the strongest model"

    oai.tried = []
    txt, _model = appmod._vimi_completion(oai, [], 500)
    assert txt == "ok"
    assert chain[0] not in oai.tried, "a retired model must not be probed again"


def test_a_transient_failure_does_not_retire_a_model(clean_model_memo):
    """Rate limits and 5xx are about this request, not the model. Disqualifying
    the strongest model over one would silently degrade every later answer."""
    chain = appmod._vimi_model_chain()
    oai = _DeadModelOAI({chain[0]}, status=429)
    appmod._vimi_completion(oai, [], 500)
    assert chain[0] not in appmod._VIMI_DEAD


def test_a_bad_api_key_does_not_retire_the_whole_chain(clean_model_memo):
    """401 fails every model identically, so treating it as permanent would turn
    a fixable credential problem into a process-long degradation."""
    chain = appmod._vimi_model_chain()
    oai = _DeadModelOAI(set(chain), status=401)
    with pytest.raises(Exception):
        appmod._vimi_completion(oai, [], 500)
    assert not appmod._VIMI_DEAD


def test_the_chain_is_never_empty_even_if_every_model_is_retired(clean_model_memo):
    appmod._VIMI_DEAD.update(appmod._vimi_model_chain(skip_dead=False))
    assert appmod._vimi_model_chain(), "an empty chain would mask the real error"


# ── Contact fields reach the answer only when asked for ──────────────────────

_ENRICHED = {"matched": True, "name": "Jane Doe", "title": "CMO",
             "email": "jane@acme.com", "apollo_email": "jane@acme.com",
             "emails": [{"email": "jane@acme.com"}],
             "phones": [{"number": "+1 555 0100"}],
             "company": {"name": "Acme"}, "location": "Austin, TX"}


def test_no_contact_field_reaches_the_answer_unless_requested():
    """_apollo_person_normalize carries FOUR keys with contact data. A denylist
    naming two of them quietly handed the model the other two on every answer."""
    facts = appmod._cpi_answer_person(_ENRICHED, wants_contact=False)
    for key in ("email", "apollo_email", "emails", "phones"):
        assert key not in facts, "%s must not reach the answer prompt" % key
    assert facts["name"] == "Jane Doe" and facts["title"] == "CMO"


def test_contact_fields_are_included_when_requested():
    facts = appmod._cpi_answer_person(_ENRICHED, wants_contact=True)
    assert facts["emails"] and facts["phones"] and facts["email"]


def test_a_new_normalizer_field_cannot_leak_by_default():
    """The allowlist has to fail closed: an unknown key is dropped, so adding a
    field to the normalizer cannot start volunteering it."""
    facts = appmod._cpi_answer_person(dict(_ENRICHED, home_address="12 Elm St"),
                                      wants_contact=True)
    assert "home_address" not in facts


# ── The reveal only pays for names Apollo actually withheld ──────────────────

@pytest.mark.parametrize("row,needs", [
    ({"full_name": "Jane Doe", "last_name": "Doe"}, False),
    ({"full_name": "Jane H.", "name_masked": True}, True),
    ({"full_name": "Sanjeev"}, True),          # first name only, carries no flag
    ({"full_name": ""}, True),
    ({"full_name": "Jane Doe"}, False),        # two tokens, no last_name key
])
def test_name_completeness_decides_whether_a_credit_is_spent(row, needs):
    assert appmod._cpi_name_incomplete(row) is needs


def test_reveal_skips_rows_that_already_have_a_full_name(no_postgres, monkeypatch):
    """A list answer used to spend one credit per person it mentioned, including
    every person whose surname had already come back free."""
    import tracker.apollo_client as ac
    calls = []
    monkeypatch.setattr(ac, "bulk_match_people",
                        lambda ids, key: calls.append(list(ids)) or {})
    people = [{"id": "p%d" % i, "full_name": "Person %d" % i, "last_name": "%d" % i}
              for i in range(10)]
    appmod._cpi_reveal_names(people, "key")
    assert calls == [], "nothing to reveal means no Apollo call at all"


def test_reveal_is_capped_and_reports_its_cost(no_postgres, monkeypatch):
    import tracker.apollo_client as ac
    seen = []

    def _bulk(ids, key):
        seen.extend(ids)
        return {i: {"id": i, "first_name": "Real", "last_name": "Name"} for i in ids}

    monkeypatch.setattr(ac, "bulk_match_people", _bulk)
    people = [{"id": "p%d" % i, "full_name": "P%d H." % i, "name_masked": True}
              for i in range(40)]
    spend = {"credits": 0}
    appmod._cpi_reveal_names(people, "key", spend=spend)
    assert len(seen) == appmod._CPI_CHAT_REVEAL_CAP
    assert len(seen) == len(set(seen)), "no id may be submitted twice"
    assert spend["credits"] == appmod._CPI_CHAT_REVEAL_CAP


def test_chat_reply_states_its_cost_and_omits_it_when_free():
    with appmod.app.test_request_context():
        assert appmod._cpi_chat_reply({"credits": 3}, answer="x").get_json()["credits"] == 3
        assert "credits" not in appmod._cpi_chat_reply({"credits": 0}, answer="x").get_json()


# ── History: one search is one entry ─────────────────────────────────────────

class _FakeCursor:
    """Records the SQL a request issues and hands back queued fetchone results."""

    def __init__(self, log, rows):
        self.log, self.rows = log, rows

    def execute(self, sql, params=None):
        self.log.append((" ".join(sql.split()), params))

    def fetchone(self):
        return self.rows.pop(0) if self.rows else None

    def fetchall(self):
        return []

    def __enter__(self):
        return self

    def __exit__(self, *a):
        return False


class _FakeConn:
    def __init__(self, rows=None):
        self.log, self.rows, self.commits = [], list(rows or []), 0

    def cursor(self):
        return _FakeCursor(self.log, self.rows)

    def commit(self):
        self.commits += 1

    def rollback(self):
        pass

    def close(self):
        pass

    def verbs(self):
        return [sql.split()[0].upper() for sql, _ in self.log
                if sql.split()[0].upper() in ("INSERT", "UPDATE", "DELETE")]


@pytest.fixture
def fake_pg(monkeypatch):
    """Swap in a recording connection and reset the one-shot DDL guard, which
    would otherwise leak table-created state between tests."""
    def _make(rows=None):
        conn = _FakeConn(rows)
        monkeypatch.setattr(appmod, "_pg_conn", lambda: conn)
        monkeypatch.setattr(appmod, "_CPI_HISTORY_TABLE_READY", False)
        return conn
    return _make


def test_paging_grows_the_existing_history_entry(fake_pg, client):
    """Without this, every Load more wrote another entry holding a superset of
    the last one, evicting real history against the 60-per-user cap."""
    import datetime as _dt
    conn = fake_pg(rows=[(7, _dt.datetime(2026, 8, 6))])   # the UPDATE finds row 7
    body = client.post("/p2/gtm/company-people-intelligence/history",
                       json={"entity": "people", "filters": {"titles": ["CMO"]},
                             "rows": [{"id": "p1"}], "replace_id": 7}).get_json()
    assert body["saved"] is True and body["id"] == 7
    assert "UPDATE" in conn.verbs()
    assert "INSERT" not in conn.verbs(), "paging must not fork a second entry"


def test_a_new_search_inserts_a_fresh_entry(fake_pg, client):
    import datetime as _dt
    conn = fake_pg(rows=[(9, _dt.datetime(2026, 8, 6))])
    body = client.post("/p2/gtm/company-people-intelligence/history",
                       json={"entity": "people", "filters": {}, "rows": [{"id": "p1"}]}).get_json()
    assert body["id"] == 9
    assert "INSERT" in conn.verbs() and "UPDATE" not in conn.verbs()


def test_an_unknown_replace_id_falls_through_to_an_insert(fake_pg, client):
    """A replace_id belonging to someone else matches nothing, and the save must
    still land under the signed-in user rather than being dropped."""
    import datetime as _dt
    conn = fake_pg(rows=[None, (11, _dt.datetime(2026, 8, 6))])
    body = client.post("/p2/gtm/company-people-intelligence/history",
                       json={"entity": "people", "filters": {}, "rows": [{"id": "p1"}],
                             "replace_id": 999999}).get_json()
    assert body["id"] == 11
    verbs = conn.verbs()
    assert verbs.count("UPDATE") == 1 and verbs.count("INSERT") == 1


def test_history_save_scopes_every_write_to_the_signed_in_user(fake_pg, client):
    """email in the WHERE clause is the authorization boundary here."""
    import datetime as _dt
    conn = fake_pg(rows=[(7, _dt.datetime(2026, 8, 6))])
    client.post("/p2/gtm/company-people-intelligence/history",
                json={"entity": "people", "filters": {}, "rows": [{"id": "p1"}],
                      "replace_id": 7})
    for sql, params in conn.log:
        verb = sql.split()[0].upper()
        if verb in ("UPDATE", "DELETE"):
            assert "reporting@position2.com" in (params or ()), sql


def test_expired_history_is_pruned_by_age_not_only_by_count(fake_pg, client):
    """These rows hold revealed emails and phone numbers, so they must not be
    retained indefinitely just because the count cap has not been reached."""
    import datetime as _dt
    conn = fake_pg(rows=[(9, _dt.datetime(2026, 8, 6))])
    client.post("/p2/gtm/company-people-intelligence/history",
                json={"entity": "people", "filters": {}, "rows": [{"id": "p1"}]})
    aged = [(sql, params) for sql, params in conn.log
            if sql.startswith("DELETE") and "make_interval" in sql]
    assert aged, "no age-based prune was issued"
    assert appmod._CPI_HISTORY_TTL_DAYS in (aged[0][1] or ())


# ── Bulk-enrich cap reporting ────────────────────────────────────────────────

def test_capped_is_only_true_when_rows_were_actually_dropped(client, no_postgres, monkeypatch):
    """Comparing the truncated length to the cap told a user who selected exactly
    50 that "only the first 50 were enriched", when nothing had been left out."""
    import tracker.apollo_client as ac
    monkeypatch.setattr(ac, "bulk_match_people",
                        lambda ids, key: {i: {"id": i} for i in ids})
    monkeypatch.setenv("APOLLO_API_KEY", "k")
    cap = appmod._CPI_BULK_ENRICH_CAP

    def _post(n):
        return client.post("/p2/gtm/company-people-intelligence/enrich-bulk",
                           json={"ids": ["id%d" % i for i in range(n)]}).get_json()

    assert _post(cap - 1)["capped"] is False
    assert _post(cap)["capped"] is False, "a full selection is not a truncated one"
    assert _post(cap + 1)["capped"] is True


def test_duplicate_ids_do_not_count_towards_the_cap(client, no_postgres, monkeypatch):
    import tracker.apollo_client as ac
    monkeypatch.setattr(ac, "bulk_match_people",
                        lambda ids, key: {i: {"id": i} for i in ids})
    monkeypatch.setenv("APOLLO_API_KEY", "k")
    ids = ["same"] * (appmod._CPI_BULK_ENRICH_CAP + 20)
    body = client.post("/p2/gtm/company-people-intelligence/enrich-bulk",
                       json={"ids": ids}).get_json()
    assert body["capped"] is False and body["fetched"] == 1


# ── Company enrichment falls back instead of dead-ending ─────────────────────

def test_company_enrichment_falls_back_from_id_to_domain(monkeypatch):
    """organizations/enrich is domain-keyed. Calling only the id form when an id
    was known, which is always, made every company profile question answer
    "Apollo doesn't have a full profile"."""
    import tracker.apollo_client as ac
    tried = []

    def _by_id(apollo_id, key):
        tried.append("id")
        return {}

    def _by_domain(domain, key):
        tried.append("domain")
        return {"id": "o1", "name": "Acme", "primary_domain": domain}

    monkeypatch.setattr(ac, "enrich_company_by_id", _by_id)
    monkeypatch.setattr(ac, "enrich_company", _by_domain)
    monkeypatch.setattr(ac, "get_leadership", lambda *a, **k: [])
    monkeypatch.setenv("APOLLO_API_KEY", "k")

    profile = appmod._cpi_enrich_company("acme.com", "o1")
    assert profile.get("matched") is True
    assert tried == ["id", "domain"], "the domain form must be reached"


def test_company_enrichment_bills_only_a_real_match(monkeypatch):
    import tracker.apollo_client as ac
    monkeypatch.setattr(ac, "enrich_company_by_id", lambda *a, **k: {})
    monkeypatch.setattr(ac, "enrich_company", lambda *a, **k: {})
    monkeypatch.setenv("APOLLO_API_KEY", "k")
    spend = {"credits": 0}
    assert appmod._cpi_enrich_company("acme.com", "o1", spend=spend)["matched"] is False
    assert spend["credits"] == 0, "a miss costs nothing and must not be counted"


# ── Apollo facts combined with research ──────────────────────────────────────

def test_the_answer_prompt_carries_both_sources(monkeypatch):
    seen = {}

    def _completion(oai, messages, max_tokens, temperature=None):
        seen["user"] = messages[-1]["content"]
        return "answer", "m"

    monkeypatch.setattr(appmod, "_vimi_completion", _completion)
    appmod._cpi_grounded_answer(None, {"company": {"name": "Acme"}}, "tell me about Acme",
                                research="Acme sells widgets.")
    assert "<apollo_facts>" in seen["user"] and "Acme" in seen["user"]
    assert "<web_research>" in seen["user"] and "widgets" in seen["user"]
    # Both blocks stay fenced and labelled as data, so third-party text inside
    # them cannot be read as instructions.
    assert "not instructions" in seen["user"]


def test_an_empty_facts_block_is_omitted_rather_than_sent_empty(monkeypatch):
    """A general question has nothing on file, and an empty facts block invites
    the model to imply we hold records we do not."""
    seen = {}

    def _completion(oai, messages, max_tokens, temperature=None):
        seen["user"] = messages[-1]["content"]
        return "answer", "m"

    monkeypatch.setattr(appmod, "_vimi_completion", _completion)
    appmod._cpi_grounded_answer(None, {}, "what is ABM?", research="ABM is...")
    assert "<apollo_facts>" not in seen["user"]
    assert "<web_research>" in seen["user"]


def test_em_dashes_never_survive_into_an_answer(monkeypatch):
    monkeypatch.setattr(appmod, "_vimi_completion",
                        lambda o, m, t, temperature=None: ("Acme — a widget maker", "m"))
    out = appmod._cpi_grounded_answer(None, {}, "q", research="r")
    assert "—" not in out
    assert " , " not in out, "the replacement must not leave a space before the comma"


def test_research_degrades_to_model_knowledge_without_a_web_tool(monkeypatch):
    """No web-search tool on the key must not silently remove the research half of
    every answer."""
    monkeypatch.setattr(appmod, "_responses_web_search", lambda *a, **k: (None, False))
    monkeypatch.setattr(appmod, "_vimi_completion",
                        lambda o, m, t, temperature=None: ("brief", "m"))
    text, used_web = appmod._cpi_research(None, "tell me about Acme")
    assert text == "brief" and used_web is False


def test_research_failure_leaves_the_apollo_answer_intact(monkeypatch):
    """Research is an enhancement. If it cannot run at all, the answer still gets
    made from the facts rather than the request failing."""
    monkeypatch.setattr(appmod, "_responses_web_search", lambda *a, **k: (None, False))

    def _boom(*a, **k):
        raise RuntimeError("no model")

    monkeypatch.setattr(appmod, "_vimi_completion", _boom)
    assert appmod._cpi_research(None, "q") == ("", False)


@pytest.mark.parametrize("org,expected_in", [
    ({"name": "Acme", "primary_domain": "acme.com"}, "acme.com"),
    ({"name": "Acme"}, "Acme"),
    ({}, ""),
    (None, ""),
])
def test_research_is_pinned_to_the_resolved_company(org, expected_in):
    """A common company name would otherwise send the research off to a different
    business than the one Apollo resolved."""
    note = appmod._cpi_company_note(org)
    assert expected_in in note
    if not expected_in:
        assert note == ""


# ── Web-search availability is detected, memoised and reported ───────────────

class _WebSearchOAI:
    """Responses-API stand-in. `fail_with` is raised for every tool type tried."""

    def __init__(self, works_with=None, fail_with=None):
        self.works_with, self.fail_with, self.tried = works_with, fail_with, []
        outer = self

        def create(model, tools, input, max_output_tokens):
            tt = tools[0]["type"]
            outer.tried.append(tt)
            if outer.fail_with is not None:
                raise outer.fail_with
            if tt != outer.works_with:
                raise RuntimeError("Unknown parameter: tools[0].type")
            return types.SimpleNamespace(output_text="researched text")

        self.responses = types.SimpleNamespace(create=create)


@pytest.fixture
def clean_web_memo():
    appmod._WEB_SEARCH_OK = None
    appmod._WEB_SEARCH_TOOL = None
    yield
    appmod._WEB_SEARCH_OK = None
    appmod._WEB_SEARCH_TOOL = None


def test_a_working_web_search_is_detected_and_remembered(clean_web_memo):
    oai = _WebSearchOAI(works_with="web_search_preview")
    txt, used = appmod._responses_web_search(oai, "m", [], 500)
    assert txt == "researched text" and used is True
    assert appmod._WEB_SEARCH_OK is True

    oai.tried = []
    appmod._responses_web_search(oai, "m", [], 500)
    assert oai.tried == ["web_search_preview"], "the proven tool name should be reused"


def test_an_unavailable_web_search_stops_being_retried(clean_web_memo):
    """Otherwise every research call pays two failed round trips, forever."""
    oai = _WebSearchOAI(fail_with=RuntimeError("Unknown parameter: tools[0].type"))
    txt, used = appmod._responses_web_search(oai, "m", [], 500)
    assert txt is None and used is False
    assert appmod._WEB_SEARCH_OK is False

    oai.tried = []
    appmod._responses_web_search(oai, "m", [], 500)
    assert oai.tried == [], "a key without web search must not be probed again"


def test_an_old_sdk_without_responses_is_treated_as_unsupported(clean_web_memo):
    oai = _WebSearchOAI(fail_with=AttributeError("'OpenAI' object has no attribute 'responses'"))
    appmod._responses_web_search(oai, "m", [], 500)
    assert appmod._WEB_SEARCH_OK is False


def test_a_transient_failure_does_not_disable_web_search(clean_web_memo):
    """A rate limit is a bad minute, not a missing capability. Disabling the tool
    over one would silently strip live research from every later answer."""
    class _Rate(Exception):
        status_code = 429

    oai = _WebSearchOAI(fail_with=_Rate("slow down"))
    appmod._responses_web_search(oai, "m", [], 500)
    assert appmod._WEB_SEARCH_OK is None, "must stay undecided, not latch to False"


def test_research_reports_whether_the_web_was_actually_used(clean_web_memo, monkeypatch):
    """The reply's web_search flag is what tells anyone, including the user, that
    live web search works. It must not be true when research came from the model."""
    monkeypatch.setattr(appmod, "_responses_web_search", lambda *a, **k: (None, False))
    monkeypatch.setattr(appmod, "_vimi_completion",
                        lambda o, m, t, temperature=None: ("brief", "m"))
    text, used_web = appmod._cpi_research(None, "q")
    assert text == "brief" and used_web is False

    monkeypatch.setattr(appmod, "_responses_web_search", lambda *a, **k: ("live", True))
    text, used_web = appmod._cpi_research(None, "q")
    assert text == "live" and used_web is True
